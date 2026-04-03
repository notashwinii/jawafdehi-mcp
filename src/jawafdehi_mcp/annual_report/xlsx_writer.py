from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .constants import CANONICAL_COLUMN_ORDER
from .fs import ensure_dir

HEADER_FILL = PatternFill("solid", fgColor="DDEEFF")
REVIEW_FILL = PatternFill("solid", fgColor="FFEECC")
HEADER_FONT = Font(name="Kalimati", bold=True, size=11)
BODY_FONT = Font(name="Kalimati", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMN_WIDTHS = {
    "क्र_सं": 6,
    "उजुरीको_व्यहोरा": 45,
    "अनुसन्धानबाट_पुष्टि": 45,
    "आयोगको_निर्णय_मिति": 15,
    "आरोपपत्र_दायर_मिति": 15,
    "मुद्दा_नं": 12,
    "प्रतिवादी_सङ्ख्या": 8,
    "प्रतिवादीको_नाम": 35,
    "पद_र_कार्यालय": 40,
    "कसुर_दफा": 25,
    "बिगो_रकम": 18,
    "NEEDS_REVIEW": 14,
}


def write_section_xlsx(records: list[dict[str, object]], output_path: Path) -> None:
    ensure_dir(output_path.parent)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = output_path.stem[:31]

    columns = active_columns(records)
    for col_idx, column_name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        sheet.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS.get(column_name, 20)

    for row_idx, record in enumerate(records, start=2):
        needs_review = bool(record.get("_needs_review"))
        for col_idx, column_name in enumerate(columns, start=1):
            value = "YES" if column_name == "NEEDS_REVIEW" and needs_review else record.get(column_name)
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            if needs_review:
                cell.fill = REVIEW_FILL

    workbook.save(output_path)


def active_columns(records: list[dict[str, object]]) -> list[str]:
    columns = [
        column
        for column in CANONICAL_COLUMN_ORDER
        if any(record.get(column) not in (None, "", "None", "null") for record in records)
    ]
    if any(record.get("_needs_review") for record in records):
        columns.append("NEEDS_REVIEW")
    return columns
